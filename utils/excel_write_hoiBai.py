from openpyxl import Workbook, load_workbook
import os

def write_test_hoiBai(file_path, run_id, khoi, mon, cauhoi, expected, actual, status):
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Run ID", "Khối", "Môn", "Câu hỏi", "Expected", "Actual", "Status"])

    ws.append([run_id, khoi, mon, cauhoi, expected, actual, status])
    wb.save(file_path)
