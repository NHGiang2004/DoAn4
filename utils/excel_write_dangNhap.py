
import os
from openpyxl import Workbook, load_workbook

def write_test_dangNhap(file_path, run_id, username, password, expected, actual, status):
    """
    Ghi kết quả test vào file Excel.
    """
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        # Thêm tiêu đề
        ws.append(["Run ID", "Username", "Password", "Expected", "Actual", "Status"])

    ws.append([run_id, username, password, expected, actual[:200], status])  # cắt actual dài quá
    wb.save(file_path)
