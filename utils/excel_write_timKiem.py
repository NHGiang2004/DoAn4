# utils/excel_write_timKiem.py
import os
from openpyxl import Workbook, load_workbook

def write_test_result(file_path, run_id, keyword, expected, actual, status):
    """
    Ghi kết quả test vào file Excel.
    """
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        # Thêm tiêu đề
        ws.append(["Run ID", "Keyword", "Expected", "Actual", "Status"])

    ws.append([run_id, keyword, expected, actual[:200], status])  # cắt actual dài quá
    wb.save(file_path)
