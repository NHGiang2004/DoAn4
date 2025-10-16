import os
from openpyxl import Workbook, load_workbook

def write_test_dangKy(file_path, run_id, hoten, tendn, sdt, email, matkhau, expected, actual, status):
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Run ID", "HoVaTen", "TenDN", "SDT", "Email", "MatKhau", "Expected", "Actual", "Status"])

    ws.append([run_id, hoten, tendn, sdt, email, matkhau, expected, actual[:200], status])
    wb.save(file_path)
