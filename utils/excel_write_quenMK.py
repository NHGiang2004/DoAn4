from openpyxl import Workbook, load_workbook
import os

def write_test_quenMK(file_path, run_id, mail, expected, actual, status):
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Run ID", "Mail", "Expected", "Actual", "Status"])

    ws.append([run_id, mail, expected, actual, status])
    wb.save(file_path)
