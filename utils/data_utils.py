import csv
import json
import openpyxl
from typing import List, Dict, Any, Optional, Union
from pathlib import Path

# Đọc file CSV
def load_csv_data(filepath: str, encoding: str = 'utf-8') -> List[Dict[str, Any]]:
    """
    Đọc dữ liệu từ file CSV và trả về danh sách dictionary.
    Tự loại bỏ BOM, bỏ dòng trống, và kiểm tra lỗi header.
    """
    file_path = Path(filepath)
    if not file_path.is_file():
        raise FileNotFoundError(f"CSV file not found: {file_path}")

    try:
        # Dùng 'utf-8-sig' để loại BOM luôn (không cần mở hai lần)
        with open(file_path, newline='', encoding='utf-8-sig') as csvfile:
            reader = csv.DictReader(csvfile)

            # Kiểm tra header
            fieldnames = reader.fieldnames
            if not fieldnames or any(h is None or h.strip() == '' for h in fieldnames):
                raise ValueError("CSV file missing or invalid headers.")

            # Đọc tất cả dòng dữ liệu
            data = []
            for row in reader:
                clean_row = {k.strip(): (v.strip() if isinstance(v, str) else v)
                             for k, v in row.items() if k is not None}
                if any(value not in (None, '') for value in clean_row.values()):
                    data.append(clean_row)

            if not data:
                raise ValueError("CSV file contains no valid data rows.")

            return data

    except Exception as e:
        raise Exception(f"Error reading CSV file '{filepath}': {e}")

# Đọc file JSON
def load_json_data(filepath: str, encoding: str = 'utf-8') -> Union[List[Dict[str, Any]], Dict[str, Any]]:
    """
    Đọc dữ liệu từ file JSON và trả về dữ liệu đã parse.

    Args:
        filepath (str): Đường dẫn đến file JSON.
        encoding (str, optional): Mã hóa file, mặc định là 'utf-8'.

    Returns:
        Union[List[Dict[str, Any]], Dict[str, Any]]: Dữ liệu JSON dưới dạng list của dict hoặc dict.

    Raises:
        FileNotFoundError: Nếu file không tồn tại.
        json.JSONDecodeError: Nếu file JSON không hợp lệ.
        UnicodeDecodeError: Nếu encoding không phù hợp.
        ValueError: Nếu file rỗng hoặc không chứa dữ liệu hợp lệ.
    """
    try:
        file_path = Path(filepath)
        if not file_path.is_file():
            raise FileNotFoundError(f"JSON file not found at: {file_path}")
        
        with open(file_path, encoding=encoding) as jsonfile:
            data = json.load(jsonfile)
            if data is None or (isinstance(data, (list, dict)) and not data):
                raise ValueError("JSON file contains no valid data.")
            
            return data
    
    except FileNotFoundError:
        raise FileNotFoundError(f"JSON file not found at: {filepath}")
    except UnicodeDecodeError as e:
        raise UnicodeDecodeError(f"Invalid encoding for JSON file: {str(e)}", e.object, e.start, e.end, e.reason)
    except json.JSONDecodeError as e:
        raise json.JSONDecodeError(f"Invalid JSON format: {str(e)}", e.doc, e.pos)
    except Exception as e:
        raise Exception(f"Unexpected error while reading JSON file: {str(e)}")

# Đọc file Excel (.xlsx)
def load_excel_data(filepath: str, sheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
    """
    Đọc dữ liệu từ file Excel (.xlsx) và trả về danh sách các dictionary.
    Mỗi dictionary đại diện cho một hàng dữ liệu với key là tên cột.
    """
    try:
        file_path = Path(filepath)
        if not file_path.is_file():
            raise FileNotFoundError(f"Excel file not found at: {file_path}")

        workbook = openpyxl.load_workbook(file_path)
        if not workbook.sheetnames:
            raise ValueError("No sheets found in the Excel file.")

        # Nếu không chỉ định sheet_name thì lấy sheet đầu tiên
        sheet = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook.active

        # Đọc hàng tiêu đề
        headers = [str(cell.value).strip() if cell.value else "" for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

        # Lọc bỏ header trống
        valid_headers = [h for h in headers if h]
        if not valid_headers:
            raise KeyError("Excel sheet must have at least one non-empty header cell.")

        data = []
        for row in sheet.iter_rows(min_row=2):
            row_data = {}
            for i, cell in enumerate(row):
                if i < len(valid_headers):  # Chỉ lấy đúng số cột hợp lệ
                    key = valid_headers[i]
                    row_data[key] = cell.value
            if any(row_data.values()):  # chỉ thêm hàng nếu có dữ liệu
                data.append(row_data)

        if not data:
            raise ValueError("Excel file contains no data rows.")

        return data

    except FileNotFoundError:
        raise FileNotFoundError(f"Excel file not found at: {filepath}")
    except Exception as e:
        raise Exception(f"Error reading Excel file: {str(e)}")